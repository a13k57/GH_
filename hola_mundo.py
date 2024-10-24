import openpyxl
from openpyxl import load_workbook
import os

# Ruta del archivo Excel
archivo_excel = r'C:\Users\acastellanos\OneDrive - Alpha Hardin\Documentos\1\python curs\_GH\archivo_hola_mundo.xlsx'

# Verifica si el archivo ya existecd
if os.path.exists(archivo_excel):
    # Cargar el archivo existente
    libro = load_workbook(archivo_excel)
else:
    # Crear un nuevo libro si no existe
    libro = openpyxl.Workbook()

# Seleccionar la hoja activa (la primera por defecto)
hoja = libro.active

# Escribir "hola mundo" en la celda A1
hoja['A1'] = "hola mundo_ modificado_modifcado a las 4:43"

# Guardar el archivo de nuevo
libro.save(archivo_excel)

print(f'Se ha escrito "hola mundo" en el archivo: {archivo_excel}')